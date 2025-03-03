import psycopg2
from openpyxl import Workbook

# Database connection parameters
DB_HOST = '127.0.0.1'
DB_PORT = 5432
DB_NAME = 'postgres'
DB_USER = 'postgres'
DB_PASSWORD = 'postgres'

# Connect to the database
conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    database=DB_NAME,
    user=DB_USER,
    password=DB_PASSWORD
)

# Create a cursor object
cursor = conn.cursor()

# Fetch all table names
cursor.execute("""
    SELECT table_name 
    FROM information_schema.tables 
    WHERE table_schema = 'public'
""")
tables = cursor.fetchall()

# Create a new Excel workbook
workbook = Workbook()

# Remove the default sheet created by openpyxl
workbook.remove(workbook.active)

# Iterate over each table and fetch column names and data types
for table in tables:
    table_name = table[0]
    cursor.execute(f"""
        SELECT column_name, data_type 
        FROM information_schema.columns 
        WHERE table_name = '{table_name}'
    """)
    columns = cursor.fetchall()

    # Add a new sheet with the table name
    worksheet = workbook.create_sheet(title=table_name)

    # Write column names and data types to the sheet
    worksheet.cell(row=1, column=1, value="Column Name")
    worksheet.cell(row=1, column=2, value="Data Type")

    for idx, column in enumerate(columns, start=2):
        worksheet.cell(row=idx, column=1, value=column[0])
        worksheet.cell(row=idx, column=2, value=column[1])

# Save the workbook
workbook.save('table_headers_with_data_types.xlsx')